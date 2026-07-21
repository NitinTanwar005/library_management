"""
report.py
---------
Reporting module: view all issued books and overdue books, and
export either report to CSV or Excel (.xlsx).
"""

import csv
import os
import tkinter as tk
from tkinter import messagebox, filedialog
from datetime import date
from mysql.connector import Error

import config
import theme

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

COLUMNS = (
    "issue_id", "book_id", "title", "member_id", "member_name",
    "issue_date", "due_date", "return_date", "fine", "status",
)
HEADINGS = (
    "Issue ID", "Book ID", "Title", "Member ID", "Member Name",
    "Issue Date", "Due Date", "Return Date", "Fine", "Status",
)


class ReportWindow(tk.Toplevel):
    def __init__(self, master):
        super().__init__(master)
        self.title(f"{config.APP_NAME} - Reports")
        self.geometry("1150x680")
        self.minsize(1000, 600)
        self.configure(bg=theme.BG_APP)
        theme.apply_theme(self)

        self.current_rows = []
        self.current_report_name = "all_issued_books"

        self._build_ui()
        self.show_all_issued()

    # ------------------------------------------------------------
    def _build_ui(self):
        header = theme.make_header(
            self, "Reports", "View issued and overdue books; export to Excel or CSV"
        )
        header.pack(fill="x", padx=24, pady=(20, 10))

        toolbar = tk.Frame(self, bg=theme.BG_APP)
        toolbar.pack(fill="x", padx=24, pady=(0, 10))

        theme.StyledButton(
            toolbar, text="All Issued Books", kind="primary", width=16,
            command=self.show_all_issued,
        ).pack(side="left", padx=(0, 8))
        theme.StyledButton(
            toolbar, text="Currently Issued", kind="info", width=16,
            command=self.show_currently_issued,
        ).pack(side="left", padx=8)
        theme.StyledButton(
            toolbar, text="Overdue Books", kind="danger", width=14,
            command=self.show_overdue,
        ).pack(side="left", padx=8)

        theme.StyledButton(
            toolbar, text="Export to CSV", kind="success", width=14,
            command=self.export_csv,
        ).pack(side="right", padx=(8, 0))
        theme.StyledButton(
            toolbar, text="Export to Excel", kind="success", width=14,
            command=self.export_excel,
        ).pack(side="right")

        self.summary_label = tk.Label(
            self, text="", font=theme.FONT_BODY_BOLD, bg=theme.BG_APP, fg=theme.TEXT_MUTED
        )
        self.summary_label.pack(anchor="w", padx=24, pady=(0, 6))

        table_outer = tk.Frame(self, bg=theme.BG_APP)
        table_outer.pack(fill="both", expand=True, padx=24, pady=(0, 20))

        self.tree = theme.styled_treeview(table_outer, COLUMNS, height=16)
        for col, heading in zip(COLUMNS, HEADINGS):
            self.tree.heading(col, text=heading)
            self.tree.column(col, width=110, anchor="center")
        self.tree.column("title", width=200, anchor="w")
        self.tree.column("member_name", width=150, anchor="w")

    # ------------------------------------------------------------
    def _run_query(self, where_clause="", report_name="report"):
        try:
            conn = config.get_connection()
            cursor = conn.cursor()
            query = f"""
                SELECT ib.issue_id, ib.book_id, b.title, ib.member_id, m.name,
                       ib.issue_date, ib.due_date, ib.return_date, ib.fine, ib.status
                FROM issued_books ib
                JOIN books b ON ib.book_id = b.book_id
                JOIN members m ON ib.member_id = m.member_id
                {where_clause}
                ORDER BY ib.issue_date DESC
            """
            cursor.execute(query)
            rows = cursor.fetchall()
            cursor.close()
            conn.close()

            self.current_rows = rows
            self.current_report_name = report_name

            def is_overdue(row):
                status, due = row[9], row[6]
                return status == "Issued" and due < date.today()

            theme.fill_treeview(self.tree, rows, overdue_check=is_overdue)
            self.summary_label.config(text=f"{len(rows)} record(s) found.")
        except Error as e:
            messagebox.showerror("Database Error", f"Could not load report:\n{e}")

    def show_all_issued(self):
        self._run_query("", "all_issued_books")

    def show_currently_issued(self):
        self._run_query("WHERE ib.status = 'Issued'", "currently_issued_books")

    def show_overdue(self):
        self._run_query(
            "WHERE ib.status = 'Issued' AND ib.due_date < CURDATE()", "overdue_books"
        )

    # ------------------------------------------------------------
    def export_csv(self):
        if not self.current_rows:
            messagebox.showwarning("No Data", "There is no data to export.")
            return
        default_name = f"{self.current_report_name}.csv"
        path = filedialog.asksaveasfilename(
            initialdir=config.EXPORT_DIR,
            initialfile=default_name,
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv")],
        )
        if not path:
            return
        try:
            with open(path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(HEADINGS)
                writer.writerows(self.current_rows)
            messagebox.showinfo("Export Successful", f"Report exported to:\n{path}")
        except OSError as e:
            messagebox.showerror("Export Failed", f"Could not write CSV file:\n{e}")

    def export_excel(self):
        if not self.current_rows:
            messagebox.showwarning("No Data", "There is no data to export.")
            return
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror(
                "Missing Dependency",
                "openpyxl is not installed.\nInstall it with: pip install openpyxl",
            )
            return

        default_name = f"{self.current_report_name}.xlsx"
        path = filedialog.asksaveasfilename(
            initialdir=config.EXPORT_DIR,
            initialfile=default_name,
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
        )
        if not path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Report"

            header_fill = PatternFill(start_color="4338CA", end_color="4338CA", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)

            for col_idx, heading in enumerate(HEADINGS, start=1):
                cell = ws.cell(row=1, column=col_idx, value=heading)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            for row_idx, row in enumerate(self.current_rows, start=2):
                for col_idx, value in enumerate(row, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=str(value) if value is not None else "")

            for col_idx, heading in enumerate(HEADINGS, start=1):
                ws.column_dimensions[chr(64 + col_idx)].width = max(14, len(heading) + 4)

            wb.save(path)
            messagebox.showinfo("Export Successful", f"Report exported to:\n{path}")
        except OSError as e:
            messagebox.showerror("Export Failed", f"Could not write Excel file:\n{e}")


if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw()
    theme.apply_theme(root)
    ReportWindow(root)
    root.mainloop()
